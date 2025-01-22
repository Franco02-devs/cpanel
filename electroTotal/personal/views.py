from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from .forms import CustomUserCreationForm, TrabajadorCreationForm, AsistenciaForm
from .models import CustomUser, Trabajador, Asistencia, AsistenciaCompleta
from .scripts import generate_unique_username, diferenciaFecha,completarLimpiarAll,completarAsistencias,limpiarAsistenciasIncompletas,corregirFecha
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import render, get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.utils import timezone



@user_passes_test(lambda u: u.is_superuser)
def create_user_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.save()
            trabajador=Trabajador.objects.create(
                user=user,
                empleado_id=str(user.id),
                empleado_nombre=form.cleaned_data['first_name'],          
            )
            return redirect('user_creado', user_id=user.id)
    else:
        form = CustomUserCreationForm()

    return render(request, 'create_user.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def create_trabajador_view(request):
    if request.method == 'POST':
        form = TrabajadorCreationForm(request.POST)
        if form.is_valid():
            trabajador = form.save(commit=False)
            password = form.cleaned_data['password1']
            username=generate_unique_username(trabajador.empleado_nombre)
            user = CustomUser.objects.create_user(
                username=username,
                password=password
            )
            user.username=username+str(user.id)+str(int(user.is_staff))
            user.save()
            trabajador.user = user
            trabajador.empleado_id=str(user.id)
            trabajador.save()
            messages.success(request, "Trabajador creado exitosamente.")
            return redirect('trabajador_creado', trabajador_id=trabajador.id)
    else:
        form = TrabajadorCreationForm()

    return render(request, 'create_trabajador.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def trabajador_creado_view(request, trabajador_id):
    trabajador = Trabajador.objects.get(id=trabajador_id)
    return render(request, 'trabajador_creado.html', {'trabajador': trabajador})

@user_passes_test(lambda u: u.is_superuser)
def user_created_view(request, user_id):
    user = CustomUser.objects.get(id=user_id)
    return render(request, 'user_creado.html', {'user': user})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def dashboard_view(request):
    return render(request, 'admin_dashboard.html')

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.user_type=="admin" or user.is_superuser :
                    return redirect('home2')
                else:
                    return redirect("home")
            else:
                messages.error(request, 'Nombre de usuario o contraseña incorrectos')
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

def home_view(request):
    if request.user.is_authenticated:
        username = request.user.username
        mensaje = f"Hola {username}!"
    else:
        mensaje = "¡Inicia sesión!"
    
    return render(request, 'home.html', {'mensaje': mensaje})
    
@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def home_view2(request):
    if request.user.is_authenticated:
        username = request.user.username
        mensaje = f"Hola {username}!"
    else:
        mensaje = "¡Inicia sesión!"
    
    return render(request, 'home2.html', {'mensaje': mensaje})

@login_required
def registrar_asistencia(request):
    if request.method == 'POST':
        form = AsistenciaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            asistencia = form.save(commit=False)
            asistencia.trabajador = form.cleaned_data.get('trabajador')
            asistencia.save()
            return redirect('confirmacion_asistencia',asistencia_id=asistencia.id)
    else:
        form = AsistenciaForm(user=request.user)
    
    return render(request, 'registrar_asistencia.html', {'form': form})

@login_required
def confirmacion_asistencia(request, asistencia_id):
    # Obtener la asistencia registrada
    asistencia = Asistencia.objects.get(id=asistencia_id)
    return render(request, 'confirmacion_asistencia.html', {'asistencia': asistencia})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def asistencia_completa_view(request):
    completarLimpiarAll()
    trabajadores = Trabajador.objects.all()
    asistencias_completas = AsistenciaCompleta.objects.all()

    # Búsqueda por nombre del trabajador
    if 'search' in request.GET:
        search_query = request.GET['search']
        asistencias_completas = asistencias_completas.filter(trabajador__empleado_nombre__icontains=search_query)

    context = {
        'trabajadores': trabajadores,
        'asistencias_completas': asistencias_completas,
    }
    return render(request, 'asistencia_completa_list.html', context)

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def lista_trabajadores(request):
    trabajadores = Trabajador.objects.all()  # Obtener todos los trabajadores
    trabajadores_match = Trabajador.objects.all()  # Obtener todos los trabajadores
    if 'search' in request.GET:
        print("debug")
        search_query = request.GET['search']
        trabajadores_match = trabajadores.filter(empleado_nombre__icontains=search_query)
    context = {
        'trabajadores': trabajadores,
        'trabajadores_match': trabajadores_match,
    }
    return render(request, 'lista_trabajadores.html', context)

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def asistencias_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)  # Obtener el trabajador
    completarAsistencias(trabajador)
    limpiarAsistenciasIncompletas(trabajador)
    asistencias = Asistencia.objects.filter(trabajador=trabajador).order_by('fecha')  # Obtener sus asistencias
    return render(request, 'asistencias_trabajador.html', {'trabajador': trabajador, 'asistencias': asistencias})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def asistencia_detalle(request, id):
    # Obtenemos la asistencia por ID
    asistencia = get_object_or_404(Asistencia, id=id)
    trabajador = asistencia.trabajador  # Obtenemos el trabajador relacionado con la asistencia
    
    # Si se hace una solicitud POST para ajustar errores, llamamos a la función
    if request.method == "POST" and "corregirFecha" in request.POST:
        # Llamamos a la función que ya creaste para ajustar la asistencia
        corregirFecha(asistencia=asistencia)  # Suponiendo que tienes una función llamada ajustar_errores en el modelo Asistencia
        # Redirigimos después de ajustar
        return redirect('asistencias_trabajador', trabajador_id=trabajador.id)

    # Renderizamos la plantilla con los detalles de la asistencia
    return render(request, 'asistencia_detalle.html', {'asistencia': asistencia, 'trabajador': trabajador})

@user_passes_test(lambda u: u.is_superuser or (u.user_type=="admin"))
def generar_excel_asistencias(request):
    completarLimpiarAll()
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # Obtener todos los trabajadores
    trabajadores = Trabajador.objects.all()

    for trabajador in trabajadores:
        # Crear una hoja para cada trabajador
        hoja = wb.create_sheet(title=trabajador.empleado_nombre)

        # Encabezados
        hoja['A1'] = 'Día entrada'
        hoja['B1'] = 'Fecha entrada'
        hoja['C1'] = 'Ingreso'
        hoja['D1'] = 'Salida'
        hoja['E1'] = 'Refrigerio y/o almuerzo'
        hoja['F1'] = 'Hora Total'
        hoja['G1'] = 'Horas Ideal'
        
        bold_font = Font(bold=True)
        for cell in hoja["1:1"]:  # Fila 1
            cell.font = bold_font

        # Obtener las AsistenciaCompleta del trabajador
        asistencia_completa = AsistenciaCompleta.objects.filter(trabajador=trabajador).order_by('entrada__finalFecha')
        
        fila = 2  # Comienza en la fila 2
        for asistencia in asistencia_completa:
            ingreso = asistencia.entrada.finalFecha
            salida = asistencia.salida.finalFecha if asistencia.salida else None
            ingreso = timezone.localtime(ingreso) if ingreso else None
            salida = timezone.localtime(salida) if salida else None

            if salida:  # Si existe salida
                # Usar la función de diferencia_fecha para calcular la duración
                hora_total = diferenciaFecha(ingreso, salida)
                
                # Puedes ajustar los valores del refrigerio si es necesario
                refrigerio = "00:00:00"  # Si tienes cálculo de refrigerio, lo ajustas aquí
                dias_espanol = {
                    'Monday': 'lunes',
                    'Tuesday': 'martes',
                    'Wednesday': 'miercoles',
                    'Thursday': 'jueves',
                    'Friday': 'viernes',
                    'Saturday': 'sábado',
                    'Sunday': 'domingo'
                }

                hoja[f'A{fila}'] = dias_espanol[ingreso.strftime('%A')]
                hoja[f'B{fila}'] = ingreso.strftime('%d/%m/%Y')
                hoja[f'C{fila}'] = str(ingreso.strftime("%H:%M:%S"))
                hoja[f'D{fila}'] = str(salida.strftime("%H:%M:%S"))
                hoja[f'E{fila}'] = refrigerio
                hoja[f'F{fila}'] = str(hora_total)
                hoja[f'G{fila}'] = None

                fila += 1

        # Ajustar ancho de las columnas
        for col in range(1, 7):
            hoja.column_dimensions[get_column_letter(col)].width = 30

        # Centrar el texto en todas las celdas
        for row in hoja.iter_rows(min_row=1, max_row=fila - 1, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # Guardar como respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Control de horario.xlsx'
    wb.save(response)
    return response







